"""
Smart Export Engine — Sistem Monitoring Capex PT Dahana
=======================================================
Memuat Template Monitoring Capex-R2.xlsx sebagai master template lalu menginjeksi
data dari database ke sel yang tepat tanpa menghapus style, warna, font, atau
merged cells yang sudah ada di template asli.
"""

import os
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import MergedCell

from ..core.config import get_settings
from ..core.database import get_supabase_admin

BULAN_COLS = ["Januari", "Februari", "Maret", "April", "Mei", "Juni",
              "Juli", "Agustus", "September", "Oktober", "November", "Desember"]

SHEET_CONFIG = {
    "RKAP": { "header_row": 1, "data_start_row": 3 },
    "Real": { "header_row": 1, "data_start_row": 3 },
    "PO": { "header_row": 4, "data_start_row": 5 },
    "Tender": { "header_row": 4, "data_start_row": 5 },
    "Kajian": { "header_row": 4, "data_start_row": 5 },
    "BAADK": { "header_row": 4, "data_start_row": 5 },
    "Lainnya": { "header_row": 4, "data_start_row": 5 },
    "Timeline": { "header_row": 4, "data_start_row": 6 },
    "LKU": { "header_row": 4, "data_start_row": 5 },
    "Data Aset": { "header_row": 4, "data_start_row": 5 },
}

def _scan_header_columns(ws: Worksheet, header_row: int) -> dict[str, int]:
    col_map: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=header_row, column=col_idx).value
        if cell_val and isinstance(cell_val, str):
            val = cell_val.strip()
            # Keterangan can appear twice in status sheets
            if val == "Keterangan" and "Keterangan" in col_map:
                col_map["Keterangan_Rekap"] = col_idx
            else:
                col_map[val] = col_idx
    return col_map

def _format_rupiah(value: int | None) -> int:
    return value or 0

def _set_value_safely(ws: Worksheet, row: int, col: int, value: any) -> None:
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_col <= col <= merged_range.max_col and merged_range.min_row <= row <= merged_range.max_row:
                ws.cell(row=merged_range.min_row, column=merged_range.min_col).value = value
                break
    else:
        cell.value = value

def _inject_real_sheet(ws: Worksheet, tahun: int, col_map: dict[str, int], start_row: int) -> None:
    client = get_supabase_admin()
    masters = client.table("capex_master").select("*").eq("tahun", tahun).order("kode").execute().data
    reals = client.table("capex_realization").select("*").eq("tahun", tahun).execute().data
    
    real_map = {(r["capex_id"], r["bulan"]): r for r in reals}
    row = start_row
    idx = 1
    
    for m in masters:
        cid = str(m["id"])
        if "No" in col_map: _set_value_safely(ws, row, col_map["No"], idx)
        if "Tahun" in col_map: _set_value_safely(ws, row, col_map["Tahun"], m.get("tahun"))
        if "Kode" in col_map: _set_value_safely(ws, row, col_map["Kode"], m.get("kode"))
        if "Kategori" in col_map: _set_value_safely(ws, row, col_map["Kategori"], m.get("kategori"))
        if "Daftar Capex" in col_map: _set_value_safely(ws, row, col_map["Daftar Capex"], m.get("daftar_capex"))
        if "Anggaran" in col_map:
            _set_value_safely(ws, row, col_map["Anggaran"], _format_rupiah(m.get("anggaran_rkap")))
            _set_value_safely(ws, row, col_map["Anggaran"] + 1, _format_rupiah(m.get("anggaran_perubahan")))
        if "PIC" in col_map: _set_value_safely(ws, row, col_map["PIC"], m.get("pic"))

        total_realisasi = 0
        total_rkap_monthly = 0
        for b in range(1, 13):
            b_name = BULAN_COLS[b - 1]
            entry = real_map.get((cid, b), {})
            rkap_m = entry.get("nilai_rkap", 0)
            real_m = entry.get("nilai_realisasi", 0)
            total_rkap_monthly += rkap_m
            total_realisasi += real_m
            
            if b_name in col_map:
                _set_value_safely(ws, row, col_map[b_name], _format_rupiah(rkap_m))
                _set_value_safely(ws, row, col_map[b_name] + 1, _format_rupiah(real_m))

        status_entry = next((r for r in reals if r["capex_id"] == cid and r.get("status")), {})
        if "Status" in col_map: _set_value_safely(ws, row, col_map["Status"], status_entry.get("status"))
        if "Keterangan" in col_map: _set_value_safely(ws, row, col_map["Keterangan"], status_entry.get("keterangan"))
        
        if "Total" in col_map:
            # We match the web logic: Total RKAP is the capex budget, Total Realisasi is sum of months
            # If the excel expects sum of monthly RKAP, we can use total_rkap_monthly, 
            # but web uses anggaran_rkap. Let's use m.get("anggaran_rkap") for the first Total column,
            # or if it's "Anggaran Aktif" it would be anggaran_perubahan.
            # In the image, Total is RKAP | Realisasi. We will use anggaran_rkap.
            _set_value_safely(ws, row, col_map["Total"], _format_rupiah(m.get("anggaran_rkap")))
            _set_value_safely(ws, row, col_map["Total"] + 1, _format_rupiah(total_realisasi))
            
        row += 1
        idx += 1

def _inject_rkap_sheet(ws: Worksheet, tahun: int, col_map: dict[str, int], start_row: int) -> None:
    client = get_supabase_admin()
    masters = client.table("capex_master").select("*").eq("tahun", tahun).order("kode").execute().data
    reals = client.table("capex_realization").select("*").eq("tahun", tahun).execute().data
    
    real_map = {(r["capex_id"], r["bulan"]): r for r in reals}
    row = start_row
    idx = 1
    
    for m in masters:
        cid = str(m["id"])
        if "No" in col_map: _set_value_safely(ws, row, col_map["No"], idx)
        if "Tahun" in col_map: _set_value_safely(ws, row, col_map["Tahun"], m.get("tahun"))
        if "Kode" in col_map: _set_value_safely(ws, row, col_map["Kode"], m.get("kode"))
        if "Kategori" in col_map: _set_value_safely(ws, row, col_map["Kategori"], m.get("kategori"))
        if "PIC" in col_map: _set_value_safely(ws, row, col_map["PIC"], m.get("pic"))
        if "Daftar Capex" in col_map: _set_value_safely(ws, row, col_map["Daftar Capex"], m.get("daftar_capex"))
        
        status_entry = next((r for r in reals if r["capex_id"] == cid and r.get("status")), {})
        if "Status" in col_map: _set_value_safely(ws, row, col_map["Status"], status_entry.get("status"))

        total_realisasi = 0
        total_rkap_monthly = 0
        for b in range(1, 13):
            b_name = BULAN_COLS[b - 1]
            entry = real_map.get((cid, b), {})
            rkap_m = entry.get("nilai_rkap", 0)
            real_m = entry.get("nilai_realisasi", 0)
            total_rkap_monthly += rkap_m
            total_realisasi += real_m
            
            if b_name in col_map:
                _set_value_safely(ws, row, col_map[b_name], _format_rupiah(rkap_m))
                _set_value_safely(ws, row, col_map[b_name] + 1, _format_rupiah(real_m))

        if "Total" in col_map:
            _set_value_safely(ws, row, col_map["Total"], _format_rupiah(m.get("anggaran_rkap")))
            _set_value_safely(ws, row, col_map["Total"] + 1, _format_rupiah(m.get("anggaran_perubahan")))
            _set_value_safely(ws, row, col_map["Total"] + 2, _format_rupiah(total_realisasi))
            
        row += 1
        idx += 1

def _inject_status_sheet(ws: Worksheet, tahun: int, status_type: str, col_map: dict[str, int], start_row: int) -> None:
    client = get_supabase_admin()
    masters = client.table("capex_master").select("*").eq("tahun", tahun).order("kode").execute().data
    status_logs = client.table("capex_status_log").select("*").eq("tahun", tahun).eq("status_type", status_type).execute().data
    
    log_map = {s["capex_id"]: s for s in status_logs}
    row = start_row
    
    for idx, m in enumerate(masters, start=1):
        cid = str(m["id"])
        log = log_map.get(cid, {})
        
        if "No" in col_map: _set_value_safely(ws, row, col_map["No"], idx)
        if "Daftar Capex" in col_map: _set_value_safely(ws, row, col_map["Daftar Capex"], m.get("daftar_capex"))
        if "Anggaran" in col_map:
            _set_value_safely(ws, row, col_map["Anggaran"], _format_rupiah(log.get("anggaran_rkap", 0)))
            _set_value_safely(ws, row, col_map["Anggaran"] + 1, _format_rupiah(log.get("anggaran_perubahan", 0)))
        
        realisasi_col = f"Realisasi {status_type}"
        if realisasi_col in col_map:
            _set_value_safely(ws, row, col_map[realisasi_col], _format_rupiah(log.get("total_realisasi", 0)))
            
        ket_full = log.get("keterangan", "") or ""
        ket_main = ket_full
        ket_rekap = ""
        if "|||" in ket_full:
            parts = ket_full.split("|||")
            ket_main = parts[0]
            ket_rekap = parts[1] if len(parts) > 1 else ""
            
        if "Keterangan" in col_map: _set_value_safely(ws, row, col_map["Keterangan"], ket_main)
        if "Keterangan_Rekap" in col_map: _set_value_safely(ws, row, col_map["Keterangan_Rekap"], ket_rekap)
        elif "Rekap Nilai" in col_map:
            _set_value_safely(ws, row, col_map["Rekap Nilai"], ket_rekap)
            _set_value_safely(ws, row, col_map["Rekap Nilai"] + 1, _format_rupiah(log.get("rekap_nilai", 0)))
            
        row += 1

def _inject_timeline_sheet(ws: Worksheet, tahun: int, col_map: dict[str, int], start_row: int) -> None:
    client = get_supabase_admin()
    masters = client.table("capex_master").select("*").eq("tahun", tahun).order("kode").execute().data
    timelines = client.table("capex_timeline").select("*").eq("tahun", tahun).execute().data
    
    time_map = {}
    for t in timelines:
        cid = str(t["capex_id"])
        if cid not in time_map: time_map[cid] = {}
        if t["bulan"] not in time_map[cid]: time_map[cid][t["bulan"]] = {}
        time_map[cid][t["bulan"]][t["minggu"]] = t["kode_status"]
        
    row = start_row
    for idx, m in enumerate(masters, start=1):
        cid = str(m["id"])
        tm = time_map.get(cid, {})
        
        if "No" in col_map: _set_value_safely(ws, row, col_map["No"], idx)
        if "Daftar Capex" in col_map: _set_value_safely(ws, row, col_map["Daftar Capex"], m.get("daftar_capex"))
        if "Nilai" in col_map: _set_value_safely(ws, row, col_map["Nilai"], _format_rupiah(m.get("anggaran_rkap", 0)))
        
        for b in range(1, 13):
            b_name = BULAN_COLS[b - 1]
            if b_name in col_map:
                base_col = col_map[b_name]
                for minggu in range(1, 5):
                    status = tm.get(b, {}).get(minggu, "")
                    _set_value_safely(ws, row, base_col + minggu - 1, status)
        row += 1

def _inject_lku_sheet(ws: Worksheet, tahun: int, col_map: dict[str, int], start_row: int) -> None:
    client = get_supabase_admin()
    lkus = client.table("capex_lku").select("*, capex_master(daftar_capex)").eq("tahun", tahun).execute().data
    
    row = start_row
    for idx, lku in enumerate(lkus, start=1):
        if "Uraian" in col_map:
            uraian = lku.get("capex_master", {}).get("daftar_capex") if lku.get("capex_master") else ""
            _set_value_safely(ws, row, col_map["Uraian"], uraian)
        if "Departemen" in col_map: _set_value_safely(ws, row, col_map["Departemen"], lku.get("departemen"))
        if "RKAP Nilai" in col_map: _set_value_safely(ws, row, col_map["RKAP Nilai"], _format_rupiah(lku.get("rkap_nilai")))
        if "RKAP Target" in col_map: _set_value_safely(ws, row, col_map["RKAP Target"], _format_rupiah(lku.get("rkap_target")))
        if "Rencana TWI" in col_map: _set_value_safely(ws, row, col_map["Rencana TWI"], _format_rupiah(lku.get("rencana_twi")))
        if "Realisasi PO" in col_map: _set_value_safely(ws, row, col_map["Realisasi PO"], _format_rupiah(lku.get("realisasi_po")))
        if "Realisasi BAST" in col_map: _set_value_safely(ws, row, col_map["Realisasi BAST"], _format_rupiah(lku.get("realisasi_bast")))
        
        rencana = lku.get("rencana_per_bulan", {}) or {}
        for b in range(1, 13):
            b_name = BULAN_COLS[b - 1]
            if b_name in col_map:
                val = rencana.get(str(b), 0)
                _set_value_safely(ws, row, col_map[b_name], _format_rupiah(val))
        row += 1

def _inject_aset_sheet(ws: Worksheet, col_map: dict[str, int], start_row: int) -> None:
    client = get_supabase_admin()
    assets = client.table("capex_assets").select("*").order("tanggal_po").execute().data

    field_map = {
        "no_po": "PO", "tanggal_po": "PO Date", "no_asset": "No. Asset", "sub_number": "SNo.",
        "category": "Category", "capitalized_on": "Cap. Date", "asset_description": "Asset Description",
        "acquis_val": "Acquis. val.", "accum_dep": "Accum. dep.", "book_val": "Book val.",
        "currency": "Crcy", "location_code": "Loc", "lokasi": "LOKASI", "room": "Room", "keterangan": "Ket"
    }

    row = start_row
    for idx, asset in enumerate(assets, start=1):
        if "No" in col_map: _set_value_safely(ws, row, col_map["No"], idx)
        for field, lbl in field_map.items():
            if lbl in col_map: _set_value_safely(ws, row, col_map[lbl], asset.get(field))
        row += 1

def generate_export(tahun: int) -> BytesIO:
    settings = get_settings()
    
    base_dir = os.path.dirname(__file__)
    possible_paths = [
        os.path.abspath(os.path.join(base_dir, "..", "..", "templates", "Template Monitoring Capex-R2.xlsx")),
        os.path.abspath(os.path.join(base_dir, "..", "..", settings.excel_template_path)),
        os.path.abspath(os.path.join(base_dir, "..", "..", "..", "docs", "Template Monitoring Capex-R2.xlsx")),
        os.path.abspath(os.path.join(os.getcwd(), "templates", "Template Monitoring Capex-R2.xlsx")),
        os.path.abspath(os.path.join(os.getcwd(), "docs", "Template Monitoring Capex-R2.xlsx")),
    ]

    template_path = None
    for p in possible_paths:
        if os.path.exists(p):
            template_path = p
            break

    if not template_path:
        raise FileNotFoundError(f"Template Excel tidak ditemukan di lokasi mana pun: {possible_paths}")

    wb = load_workbook(template_path)

    for sheet_name, config in SHEET_CONFIG.items():
        if sheet_name not in wb.sheetnames:
            continue

        ws: Worksheet = wb[sheet_name]
        col_map = _scan_header_columns(ws, config["header_row"])
        start_row = config["data_start_row"]

        if sheet_name == "Real": _inject_real_sheet(ws, tahun, col_map, start_row)
        elif sheet_name == "RKAP": _inject_rkap_sheet(ws, tahun, col_map, start_row)
        elif sheet_name in ["PO", "Tender", "Kajian", "BAADK", "Lainnya"]:
            _inject_status_sheet(ws, tahun, sheet_name, col_map, start_row)
        elif sheet_name == "Timeline": _inject_timeline_sheet(ws, tahun, col_map, start_row)
        elif sheet_name == "LKU": _inject_lku_sheet(ws, tahun, col_map, start_row)
        elif sheet_name == "Data Aset": _inject_aset_sheet(ws, col_map, start_row)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_ytd_summary_excel(tahun: int, bulan: int) -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from .dashboard import get_summary_table_ytd
    
    # Ambil data
    data = get_summary_table_ytd(tahun, bulan)
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Summary YTD {tahun}"
    
    bulan_names = ["Jan", "Feb", "Mar", "Apr", "Mei", "Jun", "Jul", "Agu", "Sep", "Okt", "Nov", "Des"]
    bulan_name = bulan_names[bulan - 1] if 1 <= bulan <= 12 else str(bulan)
    
    # Styles
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_font_sub = Font(color="FFFFFF", bold=True)
    header_fill_sub = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    
    kat_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    sum_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    
    bold_font = Font(bold=True)
    
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='334155'), 
        right=Side(style='thin', color='334155'), 
        top=Side(style='thin', color='334155'), 
        bottom=Side(style='thin', color='334155')
    )
    
    thin_border_body = Border(
        left=Side(style='thin', color='E2E8F0'), 
        right=Side(style='thin', color='E2E8F0'), 
        top=Side(style='thin', color='E2E8F0'), 
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    # Headers Row 1
    headers_r1 = [
        ("URAIAN", 1), 
        (f"Budget RKAP {tahun}", 1), 
        (f"RKAP sd {bulan_name} {tahun}", 1), 
        (f"Realisasi sd {bulan_name}-{str(tahun)[-2:]}", 2), 
        (f"% thd RKAP {bulan_name} {tahun}", 2)
    ]
    
    col_idx = 1
    for title, colspan in headers_r1:
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        
        if colspan > 1:
            ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + colspan - 1)
            for i in range(1, colspan):
                c = ws.cell(row=1, column=col_idx + i)
                c.fill = header_fill
                c.border = thin_border
        else:
            ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
            c2 = ws.cell(row=2, column=col_idx)
            c2.fill = header_fill
            c2.border = thin_border
            
        col_idx += colspan
        
    # Headers Row 2
    subheaders = ["By PO", "By BAST", "By PO", "By BAST"]
    start_sub = 4
    for i, title in enumerate(subheaders):
        cell = ws.cell(row=2, column=start_sub + i, value=title)
        cell.fill = header_fill_sub
        cell.font = header_font_sub
        cell.alignment = center_align
        cell.border = thin_border
        
    # Data Rows
    current_row = 3
    num_format = '#,##0'
    pct_format = '0.0%'
    
    grand_budget = 0
    grand_rkap_ytd = 0
    grand_po = 0
    grand_bast = 0
    
    for kat in data:
        # Kategori Row
        kat_cell = ws.cell(row=current_row, column=1, value=kat["kategori"])
        kat_cell.font = bold_font
        kat_cell.fill = kat_fill
        kat_cell.alignment = left_align
        kat_cell.border = thin_border_body
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        for i in range(2, 8):
            ws.cell(row=current_row, column=i).border = thin_border_body
            ws.cell(row=current_row, column=i).fill = kat_fill
        current_row += 1
        
        # Items
        for item in kat["items"]:
            row_data = [
                item["uraian"],
                item["budget"],
                item["rkap_ytd"],
                item["real_po"],
                item["real_bast"],
                item["pct_po"] / 100 if item["rkap_ytd"] > 0 else 0,
                item["pct_bast"] / 100 if item["rkap_ytd"] > 0 else 0
            ]
            for i, val in enumerate(row_data):
                c = ws.cell(row=current_row, column=i + 1, value=val)
                c.border = thin_border_body
                if i == 0:
                    c.alignment = left_align
                else:
                    c.alignment = right_align
                    if i in [1, 2, 3, 4]:
                        c.number_format = num_format
                    elif i in [5, 6]:
                        c.number_format = pct_format
            current_row += 1
            
        # Subtotal
        sum_title = ws.cell(row=current_row, column=1, value=f"Jumlah {kat['kategori']}")
        sum_title.font = bold_font
        sum_title.fill = sum_fill
        sum_title.border = thin_border_body
        
        sum_data = [
            kat["subtotal_budget"],
            kat["subtotal_rkap_ytd"],
            kat["subtotal_real_po"],
            kat["subtotal_real_bast"],
            (kat["subtotal_real_po"] / kat["subtotal_rkap_ytd"]) if kat["subtotal_rkap_ytd"] > 0 else 0,
            (kat["subtotal_real_bast"] / kat["subtotal_rkap_ytd"]) if kat["subtotal_rkap_ytd"] > 0 else 0
        ]
        for i, val in enumerate(sum_data):
            c = ws.cell(row=current_row, column=i + 2, value=val)
            c.font = bold_font
            c.fill = sum_fill
            c.border = thin_border_body
            c.alignment = right_align
            if i in [0, 1, 2, 3]:
                c.number_format = num_format
            else:
                c.number_format = pct_format
                
        grand_budget += kat["subtotal_budget"]
        grand_rkap_ytd += kat["subtotal_rkap_ytd"]
        grand_po += kat["subtotal_real_po"]
        grand_bast += kat["subtotal_real_bast"]
        
        current_row += 1
        
    # Grand Total
    gt_title = ws.cell(row=current_row, column=1, value="Total Realisasi Aset")
    gt_title.font = header_font
    gt_title.fill = header_fill
    gt_title.border = thin_border
    
    gt_data = [
        grand_budget,
        grand_rkap_ytd,
        grand_po,
        grand_bast,
        (grand_po / grand_rkap_ytd) if grand_rkap_ytd > 0 else 0,
        (grand_bast / grand_rkap_ytd) if grand_rkap_ytd > 0 else 0
    ]
    for i, val in enumerate(gt_data):
        c = ws.cell(row=current_row, column=i + 2, value=val)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
        c.alignment = right_align
        if i in [0, 1, 2, 3]:
            c.number_format = num_format
        else:
            c.number_format = pct_format
            
    # Resize columns
    ws.column_dimensions['A'].width = 45
    for l in ['B', 'C', 'D', 'E']:
        ws.column_dimensions[l].width = 20
    for l in ['F', 'G']:
        ws.column_dimensions[l].width = 15
        
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
