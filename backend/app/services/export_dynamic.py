import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
from typing import List, Dict, Any

def _setup_worksheet(ws, title):
    ws.title = title
    # Set default row height
    ws.sheet_format.defaultRowHeight = 15

def _apply_header_style(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    cell.border = thin_border

def _apply_data_style(cell, is_rupiah=False, align="left", bold=False):
    cell.font = Font(bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    cell.border = thin_border
    if is_rupiah:
        cell.number_format = '#,##0'

def _apply_group_style(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    cell.border = thin_border

def _apply_footer_style(cell, is_rupiah=False):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    cell.alignment = Alignment(horizontal="right", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    cell.border = thin_border
    if is_rupiah:
        cell.number_format = '#,##0'

def generate_rkap_excel(data: List[Dict[str, Any]], tahun: int) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    _setup_worksheet(ws, f"RKAP Master {tahun}")

    # Title
    ws.merge_cells("A1:K1")
    ws["A1"] = f"RKAP Master {tahun}"
    ws["A1"].font = Font(bold=True, size=14)

    # Headers
    headers_row1 = ["NO", "KODE", "DAFTAR CAPEX", "PIC", "TAHUN", "STATUS"]
    months = ["JANUARI", "FEBRUARI", "MARET", "APRIL", "MEI", "JUNI", "JULI", "AGUSTUS", "SEPTEMBER", "OKTOBER", "NOVEMBER", "DESEMBER"]
    
    col_idx = 1
    for h in headers_row1:
        ws.cell(row=3, column=col_idx, value=h)
        ws.merge_cells(start_row=3, start_column=col_idx, end_row=4, end_column=col_idx)
        col_idx += 1
    
    for m in months:
        ws.cell(row=3, column=col_idx, value=m)
        ws.merge_cells(start_row=3, start_column=col_idx, end_row=4, end_column=col_idx)
        col_idx += 1
        
    ws.cell(row=3, column=col_idx, value="Total")
    ws.merge_cells(start_row=3, start_column=col_idx, end_row=3, end_column=col_idx+1)
    ws.cell(row=4, column=col_idx, value="RKAP Awal")
    ws.cell(row=4, column=col_idx+1, value="RKAP Revisi")
    
    max_col = col_idx + 1

    for r in [3, 4]:
        for c in range(1, max_col + 1):
            _apply_header_style(ws.cell(row=r, column=c))
            
    # Grouping Data
    groups = {}
    for row in data:
        cat = row.get("kategori") or ""
        if cat not in groups:
            groups[cat] = []
        groups[cat].append(row)
        
    current_row = 5
    overall_totals = [0] * (12 + 2) # 12 months + 2 totals
    
    for cat in sorted(groups.keys()):
        # Group Header
        ws.cell(row=current_row, column=1, value="")
        ws.cell(row=current_row, column=2, value=cat)
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
        for c in range(5, max_col + 1):
            ws.cell(row=current_row, column=c, value="")
            
        group_totals = [0] * (12 + 2)
        for c in range(1, max_col + 1):
            _apply_group_style(ws.cell(row=current_row, column=c))
        
        current_row += 1
        
        # Group Items
        for idx, item in enumerate(groups[cat]):
            ws.cell(row=current_row, column=1, value=idx + 1)
            ws.cell(row=current_row, column=2, value=item.get("kode") or "")
            ws.cell(row=current_row, column=3, value=item.get("daftar_capex") or "")
            ws.cell(row=current_row, column=4, value=item.get("pic") or "")
            ws.cell(row=current_row, column=5, value=item.get("tahun") or "")
            ws.cell(row=current_row, column=6, value=item.get("status") or "")
            
            c_idx = 7
            for i in range(1, 13):
                rkap = item.get(f"b{i}_rkap") or 0
                ws.cell(row=current_row, column=c_idx, value=rkap)
                group_totals[i-1] += rkap
                c_idx += 1
                
            rkap_awal = item.get("anggaran_rkap") or 0
            rkap_rev = item.get("anggaran_perubahan") or 0
            
            ws.cell(row=current_row, column=c_idx, value=rkap_awal)
            ws.cell(row=current_row, column=c_idx+1, value=rkap_rev)
            
            group_totals[-2] += rkap_awal
            group_totals[-1] += rkap_rev
            
            for c in range(1, 7):
                _apply_data_style(ws.cell(row=current_row, column=c), align="center" if c in [1,5,6] else "left")
            for c in range(7, max_col + 1):
                _apply_data_style(ws.cell(row=current_row, column=c), is_rupiah=True, align="right")
                
            current_row += 1
            
        # Write Group Totals (Update Group Header)
        # Actually Group Header was written above, but we write sums into it now
        # Wait, openpyxl cell value can be modified after merging.
        g_idx = 7
        for v in group_totals:
            ws.cell(row=current_row - len(groups[cat]) - 1, column=g_idx, value=v).number_format = '#,##0'
            ws.cell(row=current_row - len(groups[cat]) - 1, column=g_idx).alignment = Alignment(horizontal="right", vertical="center")
            overall_totals[g_idx - 7] += v
            g_idx += 1
            
    # Write Footer
    ws.cell(row=current_row, column=1, value="Total Keseluruhan")
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    
    g_idx = 7
    for v in overall_totals:
        ws.cell(row=current_row, column=g_idx, value=v)
        g_idx += 1
        
    for c in range(1, max_col + 1):
        _apply_footer_style(ws.cell(row=current_row, column=c), is_rupiah=(c >= 7))
        
    # Set Column Widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 15
    for c in range(7, max_col + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 15
        
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def generate_realization_excel(data: List[Dict[str, Any]], tahun: int) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    _setup_worksheet(ws, f"Realisasi {tahun}")

    ws.merge_cells("A1:K1")
    ws["A1"] = f"Realisasi {tahun}"
    ws["A1"].font = Font(bold=True, size=14)

    # Headers
    headers_row1 = ["NO", "DAFTAR CAPEX"]
    months = ["JANUARI", "FEBRUARI", "MARET", "APRIL", "MEI", "JUNI", "JULI", "AGUSTUS", "SEPTEMBER", "OKTOBER", "NOVEMBER", "DESEMBER"]
    
    col_idx = 1
    for h in headers_row1:
        ws.cell(row=3, column=col_idx, value=h)
        ws.merge_cells(start_row=3, start_column=col_idx, end_row=4, end_column=col_idx)
        col_idx += 1
        
    ws.cell(row=3, column=col_idx, value="ANGGARAN")
    ws.merge_cells(start_row=3, start_column=col_idx, end_row=3, end_column=col_idx+1)
    ws.cell(row=4, column=col_idx, value="RKAP")
    ws.cell(row=4, column=col_idx+1, value="PERUBAHAN")
    col_idx += 2
    
    ws.cell(row=3, column=col_idx, value="STATUS")
    ws.merge_cells(start_row=3, start_column=col_idx, end_row=4, end_column=col_idx)
    col_idx += 1
    
    ws.cell(row=3, column=col_idx, value="KETERANGAN")
    ws.merge_cells(start_row=3, start_column=col_idx, end_row=4, end_column=col_idx)
    col_idx += 1
    
    for m in months:
        ws.cell(row=3, column=col_idx, value=m)
        ws.merge_cells(start_row=3, start_column=col_idx, end_row=4, end_column=col_idx)
        col_idx += 1
        
    ws.cell(row=3, column=col_idx, value="TOTAL REALISASI PO")
    ws.merge_cells(start_row=3, start_column=col_idx, end_row=4, end_column=col_idx)
    col_idx += 1
    
    ws.cell(row=3, column=col_idx, value="TOTAL BAST")
    ws.merge_cells(start_row=3, start_column=col_idx, end_row=4, end_column=col_idx)
    col_idx += 1
        
    max_col = col_idx - 1

    for r in [3, 4]:
        for c in range(1, max_col + 1):
            _apply_header_style(ws.cell(row=r, column=c))
            
    current_row = 5
    overall_totals = [0] * (12 + 2 + 2) # 2 anggaran + 12 bulan + 2 totals
    
    for idx, item in enumerate(data):
        ws.cell(row=current_row, column=1, value=idx + 1)
        ws.cell(row=current_row, column=2, value=item.get("daftar_capex") or "")
        
        rkap = item.get("anggaran_rkap") or 0
        perubahan = item.get("anggaran_perubahan") or 0
        ws.cell(row=current_row, column=3, value=rkap)
        ws.cell(row=current_row, column=4, value=perubahan)
        
        overall_totals[0] += rkap
        overall_totals[1] += perubahan
        

        ws.cell(row=current_row, column=5, value=item.get("status") or "")
        ws.cell(row=current_row, column=6, value=item.get("keterangan") or "")
        
        c_idx = 7
        for i in range(1, 13):
            bln_real = item.get(f"b{i}_real") or 0
            ws.cell(row=current_row, column=c_idx, value=bln_real)
            
            overall_totals[2 + (i-1)] += bln_real
            c_idx += 1
            
        total_real_po = sum((item.get(f"b{i}_real") or 0) for i in range(1, 13))
        total_bast = sum((item.get(f"b{i}_bast") or 0) for i in range(1, 13))
        
        ws.cell(row=current_row, column=c_idx, value=total_real_po)
        overall_totals[2 + 12] += total_real_po
        c_idx += 1
        
        ws.cell(row=current_row, column=c_idx, value=total_bast)
        overall_totals[2 + 13] += total_bast
        c_idx += 1
            
        for c in range(1, max_col + 1):
            is_rupiah = c in [3, 4] or c >= 7
            align = "center" if c in [1, 5] else ("right" if is_rupiah else "left")
            _apply_data_style(ws.cell(row=current_row, column=c), is_rupiah=is_rupiah, align=align)
            
        current_row += 1
        
    # Write Footer
    ws.cell(row=current_row, column=1, value="Total")
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    
    ws.cell(row=current_row, column=3, value=overall_totals[0])
    ws.cell(row=current_row, column=4, value=overall_totals[1])
    ws.cell(row=current_row, column=5, value="")
    ws.cell(row=current_row, column=6, value="")
    
    g_idx = 7
    for v in overall_totals[2:]:
        ws.cell(row=current_row, column=g_idx, value=v)
        g_idx += 1
        
    for c in range(1, max_col + 1):
        _apply_footer_style(ws.cell(row=current_row, column=c), is_rupiah=(c in [3,4] or c >= 7))
        if c in [5, 6]:
             ws.cell(row=current_row, column=c).fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        
    # Set Column Widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 25
    for c in range(7, max_col + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 15
        
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def generate_gabungan_excel(data: List[Dict[str, Any]], tahun: int) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    _setup_worksheet(ws, f"RKAP vs Realisasi {tahun}")

    ws.merge_cells("A1:K1")
    ws["A1"] = f"RKAP vs Realisasi {tahun}"
    ws["A1"].font = Font(bold=True, size=14)

    # Headers
    headers_row1 = ["NO", "DAFTAR CAPEX"]
    months = ["JANUARI", "FEBRUARI", "MARET", "APRIL", "MEI", "JUNI", "JULI", "AGUSTUS", "SEPTEMBER", "OKTOBER", "NOVEMBER", "DESEMBER"]
    
    col_idx = 1
    for h in headers_row1:
        ws.cell(row=3, column=col_idx, value=h)
        ws.merge_cells(start_row=3, start_column=col_idx, end_row=4, end_column=col_idx)
        col_idx += 1
        
    ws.cell(row=3, column=col_idx, value="ANGGARAN")
    ws.merge_cells(start_row=3, start_column=col_idx, end_row=3, end_column=col_idx+1)
    ws.cell(row=4, column=col_idx, value="RKAP")
    ws.cell(row=4, column=col_idx+1, value="PERUBAHAN")
    col_idx += 2
    
    ws.cell(row=3, column=col_idx, value="STATUS")
    ws.merge_cells(start_row=3, start_column=col_idx, end_row=4, end_column=col_idx)
    col_idx += 1
    
    ws.cell(row=3, column=col_idx, value="KETERANGAN")
    ws.merge_cells(start_row=3, start_column=col_idx, end_row=4, end_column=col_idx)
    col_idx += 1
    
    for m in months:
        ws.cell(row=3, column=col_idx, value=m)
        ws.merge_cells(start_row=3, start_column=col_idx, end_row=3, end_column=col_idx+1)
        ws.cell(row=4, column=col_idx, value="RKAP")
        ws.cell(row=4, column=col_idx+1, value="REALISASI")
        col_idx += 2
        
    max_col = col_idx - 1

    for r in [3, 4]:
        for c in range(1, max_col + 1):
            _apply_header_style(ws.cell(row=r, column=c))
            
    current_row = 5
    overall_totals = [0] * (24 + 2) # 2 anggaran + 24 bulan
    
    for idx, item in enumerate(data):
        ws.cell(row=current_row, column=1, value=idx + 1)
        ws.cell(row=current_row, column=2, value=item.get("daftar_capex") or "")
        
        rkap = item.get("anggaran_rkap") or 0
        perubahan = item.get("anggaran_perubahan") or 0
        ws.cell(row=current_row, column=3, value=rkap)
        ws.cell(row=current_row, column=4, value=perubahan)
        
        overall_totals[0] += rkap
        overall_totals[1] += perubahan
        

        ws.cell(row=current_row, column=5, value=item.get("status") or "")
        ws.cell(row=current_row, column=6, value=item.get("keterangan") or "")
        
        c_idx = 7
        for i in range(1, 13):
            bln_rkap = item.get(f"b{i}_rkap") or 0
            bln_real = item.get(f"b{i}_real") or 0
            ws.cell(row=current_row, column=c_idx, value=bln_rkap)
            ws.cell(row=current_row, column=c_idx+1, value=bln_real)
            
            overall_totals[2 + (i-1)*2] += bln_rkap
            overall_totals[2 + (i-1)*2 + 1] += bln_real
            c_idx += 2
            
        for c in range(1, max_col + 1):
            is_rupiah = c in [3, 4] or c >= 7
            align = "center" if c in [1, 5] else ("right" if is_rupiah else "left")
            _apply_data_style(ws.cell(row=current_row, column=c), is_rupiah=is_rupiah, align=align)
            
        current_row += 1
        
    # Write Footer
    ws.cell(row=current_row, column=1, value="Total")
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    
    ws.cell(row=current_row, column=3, value=overall_totals[0])
    ws.cell(row=current_row, column=4, value=overall_totals[1])
    ws.cell(row=current_row, column=5, value="")
    ws.cell(row=current_row, column=6, value="")
    
    g_idx = 7
    for v in overall_totals[2:]:
        ws.cell(row=current_row, column=g_idx, value=v)
        g_idx += 1
        
    for c in range(1, max_col + 1):
        _apply_footer_style(ws.cell(row=current_row, column=c), is_rupiah=(c in [3,4] or c >= 7))
        if c in [5, 6]:
             ws.cell(row=current_row, column=c).fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        
    # Set Column Widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 25
    for c in range(7, max_col + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 15
        
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_audit_logs_excel(data: List[Dict[str, Any]], tahun: int) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    _setup_worksheet(ws, f"Riwayat Pengalihan {tahun}")

    # Title
    ws.merge_cells("A1:K1")
    ws["A1"] = f"Riwayat Pengalihan Anggaran {tahun}"
    ws["A1"].font = Font(bold=True, size=14)

    # Headers
    headers = [
        "NO", 
        "NAMA KAJIAN (SUMBER)", 
        "USULAN NILAI AWAL", 
        "NILAI RKAP", 
        "SELISIH", 
        "ANGGARAN PENGGANTI (TUJUAN)", 
        "ANGGARAN AWAL", 
        "NILAI PENGALIHAN", 
        "ND PERSETUJUAN", 
        "USER", 
        "TANGGAL & WAKTU"
    ]
    
    col_idx = 1
    for h in headers:
        cell = ws.cell(row=3, column=col_idx, value=h)
        _apply_header_style(cell)
        col_idx += 1

    # Data rows
    row_idx = 4
    for i, log in enumerate(data, start=1):
        # NO
        c = ws.cell(row=row_idx, column=1, value=i)
        _apply_data_style(c, align="center")
        
        # NAMA KAJIAN
        c = ws.cell(row=row_idx, column=2, value=log.get("source_capex_name") or "-")
        _apply_data_style(c, align="left")
        
        # USULAN NILAI AWAL
        val = log.get("source_nilai_awal") or 0
        c = ws.cell(row=row_idx, column=3, value=val)
        _apply_data_style(c, is_rupiah=True, align="right")
        
        # NILAI RKAP
        val = log.get("source_nilai_akhir") or 0
        c = ws.cell(row=row_idx, column=4, value=val)
        _apply_data_style(c, is_rupiah=True, align="right")
        
        # SELISIH
        val = (log.get("source_nilai_awal") or 0) - (log.get("source_nilai_akhir") or 0)
        c = ws.cell(row=row_idx, column=5, value=val)
        _apply_data_style(c, is_rupiah=True, align="right")
        
        # ANGGARAN PENGGANTI
        c = ws.cell(row=row_idx, column=6, value=log.get("target_capex_name") or "-")
        _apply_data_style(c, align="left")
        
        # ANGGARAN AWAL
        val = log.get("target_nilai_awal") or 0
        c = ws.cell(row=row_idx, column=7, value=val)
        _apply_data_style(c, is_rupiah=True, align="right")
        
        # NILAI PENGALIHAN
        val = log.get("target_nilai_akhir") or 0
        c = ws.cell(row=row_idx, column=8, value=val)
        _apply_data_style(c, is_rupiah=True, align="right")
        
        # ND PERSETUJUAN
        c = ws.cell(row=row_idx, column=9, value=log.get("nd_persetujuan") or "-")
        _apply_data_style(c, align="center")
        
        # USER
        c = ws.cell(row=row_idx, column=10, value=log.get("user_name") or "-")
        _apply_data_style(c, align="center")
        
        # TANGGAL & WAKTU
        dt_str = log.get("created_at")
        formatted_dt = "-"
        if dt_str:
            try:
                from datetime import datetime
                dt_obj = datetime.fromisoformat(dt_str.replace('Z', '+00:00'))
                formatted_dt = dt_obj.strftime("%Y-%m-%d %H:%M:%S")
            except:
                formatted_dt = dt_str[:19].replace("T", " ")
        c = ws.cell(row=row_idx, column=11, value=formatted_dt)
        _apply_data_style(c, align="center")
        
        row_idx += 1

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 35)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream

def generate_assets_excel(data: List[Dict[str, Any]]) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    _setup_worksheet(ws, "Data Aset")
    
    ws.merge_cells("A1:R1")
    ws["A1"] = "Data Aset PT Dahana"
    ws["A1"].font = Font(bold=True, size=14)
    
    headers = [
        "Kajian No", "Kajian Tanggal", "Kajian Perihal", "No PO", "Tanggal PO", 
        "No Asset", "Sub Number", "Kategori", "Capitalized On", "Asset Description", 
        "Acquis Val", "Accum Dep", "Book Val", "Currency", "Location Code", 
        "Lokasi", "Room", "Keterangan"
    ]
    
    for c_idx, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=c_idx, value=h)
        _apply_header_style(c)
        
    row_idx = 6
    for item in data:
        c = ws.cell(row=row_idx, column=1, value=item.get("kajian_no"))
        _apply_data_style(c)
        c = ws.cell(row=row_idx, column=2, value=item.get("kajian_tanggal"))
        _apply_data_style(c)
        c = ws.cell(row=row_idx, column=3, value=item.get("kajian_perihal"))
        _apply_data_style(c)
        c = ws.cell(row=row_idx, column=4, value=item.get("no_po"))
        _apply_data_style(c)
        c = ws.cell(row=row_idx, column=5, value=item.get("tanggal_po"))
        _apply_data_style(c)
        c = ws.cell(row=row_idx, column=6, value=item.get("no_asset"))
        _apply_data_style(c)
        c = ws.cell(row=row_idx, column=7, value=item.get("sub_number"))
        _apply_data_style(c)
        c = ws.cell(row=row_idx, column=8, value=item.get("category"))
        _apply_data_style(c)
        c = ws.cell(row=row_idx, column=9, value=item.get("capitalized_on"))
        _apply_data_style(c)
        c = ws.cell(row=row_idx, column=10, value=item.get("asset_description"))
        _apply_data_style(c)
        
        c = ws.cell(row=row_idx, column=11, value=item.get("acquis_val", 0))
        _apply_data_style(c, is_rupiah=True, align="right")
        c = ws.cell(row=row_idx, column=12, value=item.get("accum_dep", 0))
        _apply_data_style(c, is_rupiah=True, align="right")
        c = ws.cell(row=row_idx, column=13, value=item.get("book_val", 0))
        _apply_data_style(c, is_rupiah=True, align="right")
        
        c = ws.cell(row=row_idx, column=14, value=item.get("currency"))
        _apply_data_style(c)
        c = ws.cell(row=row_idx, column=15, value=item.get("location_code"))
        _apply_data_style(c)
        c = ws.cell(row=row_idx, column=16, value=item.get("lokasi"))
        _apply_data_style(c)
        c = ws.cell(row=row_idx, column=17, value=item.get("room"))
        _apply_data_style(c)
        c = ws.cell(row=row_idx, column=18, value=item.get("keterangan"))
        _apply_data_style(c)
        
        row_idx += 1
        
    for col in ws.columns:
        max_length = 0
        column = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = min((max_length + 2), 35)
        
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream

def generate_timeline_excel(data: List[Dict[str, Any]], tahun: int) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    _setup_worksheet(ws, f"Timeline {tahun}")
    
    ws.merge_cells("A1:K1")
    ws["A1"] = f"Timeline (Gantt Chart) {tahun}"
    ws["A1"].font = Font(bold=True, size=14)
    
    headers_main = ["NO", "KATEGORI", "KODE", "DAFTAR CAPEX", "PIC"]
    months = ["JANUARI", "FEBRUARI", "MARET", "APRIL", "MEI", "JUNI", "JULI", "AGUSTUS", "SEPTEMBER", "OKTOBER", "NOVEMBER", "DESEMBER"]
    
    col_idx = 1
    for h in headers_main:
        ws.cell(row=3, column=col_idx, value=h)
        ws.merge_cells(start_row=3, start_column=col_idx, end_row=4, end_column=col_idx)
        col_idx += 1
        
    for m in months:
        ws.cell(row=3, column=col_idx, value=m)
        ws.merge_cells(start_row=3, start_column=col_idx, end_row=3, end_column=col_idx+4)
        for w in range(1, 6):
            ws.cell(row=4, column=col_idx, value=str(w))
            col_idx += 1
            
    # Apply style to headers
    for r in range(3, 5):
        for c in range(1, col_idx):
            cell = ws.cell(row=r, column=c)
            # Only apply if it's not merged empty cell, but openpyxl applies to top-left
            _apply_header_style(cell)
                
    row_idx = 5
    for i, item in enumerate(data, 1):
        c = ws.cell(row=row_idx, column=1, value=i)
        _apply_data_style(c, align="center")
        c = ws.cell(row=row_idx, column=2, value=item.get("kategori"))
        _apply_data_style(c)
        c = ws.cell(row=row_idx, column=3, value=item.get("kode"))
        _apply_data_style(c)
        c = ws.cell(row=row_idx, column=4, value=item.get("daftar_capex"))
        _apply_data_style(c)
        c = ws.cell(row=row_idx, column=5, value=item.get("departemen"))
        _apply_data_style(c)
        
        c_idx = 6
        timeline_dict = item.get("timeline_dict", {})
        for month_idx in range(1, 13):
            month_data = timeline_dict.get(month_idx, {})
            for week_idx in range(1, 6):
                status = month_data.get(week_idx, "")
                c = ws.cell(row=row_idx, column=c_idx, value=status)
                _apply_data_style(c, align="center")
                c_idx += 1
                
        row_idx += 1
        
    for col in ws.columns:
        max_length = 0
        column = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.row < 3: continue
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except: pass
        if col[0].column >= 6:
            ws.column_dimensions[column].width = 5
        else:
            ws.column_dimensions[column].width = min((max_length + 2), 35)
            
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
