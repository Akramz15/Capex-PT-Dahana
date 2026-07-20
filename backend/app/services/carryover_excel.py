from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from typing import List, Dict, Any
from .export_dynamic import _setup_worksheet, _apply_header_style, _apply_data_style, _apply_footer_style

def generate_carryover_excel(data: List[Dict[str, Any]], tahun: int) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    title = f"Carry Over {tahun}"
    _setup_worksheet(ws, title[:31])

    ws.merge_cells("A1:AE1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)

    # Headers Row 3
    headers_row1 = ["NO", "URAIAN", f"Carryover {tahun-1}", "User"]
    months = ["JANUARI", "FEBRUARI", "MARET", "APRIL", "MEI", "JUNI", "JULI", "AGUSTUS", "SEPTEMBER", "OKTOBER", "NOVEMBER", "DESEMBER"]
    
    col_idx = 1
    for h in headers_row1:
        ws.cell(row=3, column=col_idx, value=h)
        ws.merge_cells(start_row=3, start_column=col_idx, end_row=4, end_column=col_idx)
        col_idx += 1
        
    for m in months:
        ws.cell(row=3, column=col_idx, value=m)
        ws.merge_cells(start_row=3, start_column=col_idx, end_row=3, end_column=col_idx+1)
        ws.cell(row=4, column=col_idx, value="BA")
        ws.cell(row=4, column=col_idx+1, value="PO")
        col_idx += 2
        
    ws.cell(row=3, column=col_idx, value="TOTAL")
    ws.merge_cells(start_row=3, start_column=col_idx, end_row=3, end_column=col_idx+1)
    ws.cell(row=4, column=col_idx, value="By PO")
    ws.cell(row=4, column=col_idx+1, value="By BA")
    col_idx += 2
    
    max_col = col_idx - 1

    for r in [3, 4]:
        for c in range(1, max_col + 1):
            _apply_header_style(ws.cell(row=r, column=c))
            
    current_row = 5
    overall_totals = [0] * (1 + 24 + 2) # Carryover + 12*2 + 2 totals
    
    for idx, item in enumerate(data):
        ws.cell(row=current_row, column=1, value=idx + 1)
        ws.cell(row=current_row, column=2, value=item.get("daftar_capex") or "")
        
        rkap = item.get("anggaran_rkap") or 0
        ws.cell(row=current_row, column=3, value=rkap)
        ws.cell(row=current_row, column=4, value=item.get("pic") or "")
        
        overall_totals[0] += rkap
        
        c_idx = 5
        total_real = 0
        total_bast = 0
        for i in range(1, 13):
            bln_bast = item.get(f"b{i}_bast") or 0
            bln_real = item.get(f"b{i}_real") or 0
            ws.cell(row=current_row, column=c_idx, value=bln_bast)
            ws.cell(row=current_row, column=c_idx+1, value=bln_real)
            
            overall_totals[1 + (i-1)*2] += bln_bast
            overall_totals[2 + (i-1)*2] += bln_real
            
            total_bast += bln_bast
            total_real += bln_real
            c_idx += 2
            
        ws.cell(row=current_row, column=c_idx, value=total_real)
        overall_totals[-2] += total_real
        c_idx += 1
        
        ws.cell(row=current_row, column=c_idx, value=total_bast)
        overall_totals[-1] += total_bast
        c_idx += 1
            
        for c in range(1, max_col + 1):
            is_rupiah = c == 3 or c >= 5
            align = "center" if c in [1, 4] else ("right" if is_rupiah else "left")
            _apply_data_style(ws.cell(row=current_row, column=c), is_rupiah=is_rupiah, align=align)
            
        current_row += 1
        
    # Write Footer
    ws.cell(row=current_row, column=1, value="Total")
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    
    ws.cell(row=current_row, column=3, value=overall_totals[0])
    ws.cell(row=current_row, column=4, value="")
    
    g_idx = 5
    for v in overall_totals[1:]:
        ws.cell(row=current_row, column=g_idx, value=v)
        g_idx += 1
        
    for c in range(1, max_col + 1):
        _apply_footer_style(ws.cell(row=current_row, column=c), is_rupiah=(c == 3 or c >= 5))
        if c == 4:
             ws.cell(row=current_row, column=c).fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        
    # Set Column Widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    for c in range(5, max_col + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 15
        
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
