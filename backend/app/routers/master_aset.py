from fastapi import APIRouter, Depends, HTTPException, Query, status, UploadFile, File
from typing import Optional, List
from uuid import UUID
from datetime import datetime
import openpyxl
import pandas as pd
from io import BytesIO
from fastapi.responses import Response

from ..core.database import get_supabase_admin
from ..core.security import get_current_user, require_admin
from ..models.master_aset import (
    AsetNomorCreate, AsetNomorUpdate, AsetNomorResponse,
    AsetLaporanCreate, AsetLaporanUpdate, AsetLaporanResponse,
    AsetDataCreate, AsetDataUpdate, AsetDataResponse
)
from ..services.audit import log_module_update

router = APIRouter(prefix="/master-aset", tags=["Master Aset"])

# -------------------------
# ASET NOMOR
# -------------------------
@router.get("/nomor", response_model=list[AsetNomorResponse])
def get_aset_nomor(_user: dict = Depends(get_current_user)):
    client = get_supabase_admin()
    all_data = []
    chunk_size = 1000
    start = 0
    while True:
        result = client.table("aset_nomor").select("*").order("created_at", desc=False).range(start, start + chunk_size - 1).execute()
        data = result.data
        if not data: break
        all_data.extend(data)
        if len(data) < chunk_size: break
        start += chunk_size
    return all_data

@router.post("/nomor", response_model=AsetNomorResponse)
def create_aset_nomor(payload: AsetNomorCreate, _admin: dict = Depends(require_admin)):
    client = get_supabase_admin()
    data = payload.model_dump(exclude_none=True)
    result = client.table("aset_nomor").insert(data).execute()
    log_module_update(client, "Permintaan Nomor Aset", _admin.get("full_name", "Admin"))
    return result.data[0]

@router.put("/nomor/{id}", response_model=AsetNomorResponse)
def update_aset_nomor(id: UUID, payload: AsetNomorUpdate, _admin: dict = Depends(require_admin)):
    client = get_supabase_admin()
    data = payload.model_dump(exclude_none=True)
    result = client.table("aset_nomor").update(data).eq("id", str(id)).execute()
    if not result.data: raise HTTPException(status_code=404, detail="Not found")
    log_module_update(client, "Permintaan Nomor Aset", _admin.get("full_name", "Admin"))
    return result.data[0]

@router.delete("/nomor/{id}")
def delete_aset_nomor(id: UUID, _admin: dict = Depends(require_admin)):
    client = get_supabase_admin()
    result = client.table("aset_nomor").delete().eq("id", str(id)).execute()
    if not result.data: raise HTTPException(status_code=404, detail="Not found")
    log_module_update(client, "Permintaan Nomor Aset", _admin.get("full_name", "Admin"))
    return {"message": "Deleted"}

# -------------------------
# ASET LAPORAN AKTIVA
# -------------------------
@router.get("/laporan", response_model=list[AsetLaporanResponse])
def get_aset_laporan(_user: dict = Depends(get_current_user)):
    client = get_supabase_admin()
    all_data = []
    chunk_size = 1000
    start = 0
    while True:
        result = client.table("aset_laporan_aktiva").select("*").order("created_at", desc=False).range(start, start + chunk_size - 1).execute()
        data = result.data
        if not data: break
        all_data.extend(data)
        if len(data) < chunk_size: break
        start += chunk_size
    return all_data

@router.post("/laporan", response_model=AsetLaporanResponse)
def create_aset_laporan(payload: AsetLaporanCreate, _admin: dict = Depends(require_admin)):
    client = get_supabase_admin()
    data = payload.model_dump(exclude_none=True)
    if data.get("capitalized_on"): data["capitalized_on"] = str(data["capitalized_on"])
    result = client.table("aset_laporan_aktiva").insert(data).execute()
    log_module_update(client, "Laporan Aktiva Tetap", _admin.get("full_name", "Admin"))
    return result.data[0]

@router.put("/laporan/{id}", response_model=AsetLaporanResponse)
def update_aset_laporan(id: UUID, payload: AsetLaporanUpdate, _admin: dict = Depends(require_admin)):
    client = get_supabase_admin()
    data = payload.model_dump(exclude_none=True)
    if data.get("capitalized_on"): data["capitalized_on"] = str(data["capitalized_on"])
    result = client.table("aset_laporan_aktiva").update(data).eq("id", str(id)).execute()
    if not result.data: raise HTTPException(status_code=404, detail="Not found")
    log_module_update(client, "Laporan Aktiva Tetap", _admin.get("full_name", "Admin"))
    return result.data[0]

@router.delete("/laporan/{id}")
def delete_aset_laporan(id: UUID, _admin: dict = Depends(require_admin)):
    client = get_supabase_admin()
    result = client.table("aset_laporan_aktiva").delete().eq("id", str(id)).execute()
    if not result.data: raise HTTPException(status_code=404, detail="Not found")
    log_module_update(client, "Laporan Aktiva Tetap", _admin.get("full_name", "Admin"))
    return {"message": "Deleted"}

# -------------------------
# ASET DATA
# -------------------------
@router.get("/data", response_model=list[AsetDataResponse])
def get_aset_data(_user: dict = Depends(get_current_user)):
    client = get_supabase_admin()
    all_data = []
    chunk_size = 1000
    start = 0
    while True:
        result = client.table("aset_data").select("*").order("created_at", desc=False).range(start, start + chunk_size - 1).execute()
        data = result.data
        if not data: break
        all_data.extend(data)
        if len(data) < chunk_size: break
        start += chunk_size
    return all_data

@router.post("/data", response_model=AsetDataResponse)
def create_aset_data(payload: AsetDataCreate, _admin: dict = Depends(require_admin)):
    client = get_supabase_admin()
    data = payload.model_dump(exclude_none=True)
    if data.get("capitalized_on"): data["capitalized_on"] = str(data["capitalized_on"])
    result = client.table("aset_data").insert(data).execute()
    log_module_update(client, "Data Aset", _admin.get("full_name", "Admin"))
    return result.data[0]

@router.put("/data/{id}", response_model=AsetDataResponse)
def update_aset_data(id: UUID, payload: AsetDataUpdate, _admin: dict = Depends(require_admin)):
    client = get_supabase_admin()
    data = payload.model_dump(exclude_none=True)
    if data.get("capitalized_on"): data["capitalized_on"] = str(data["capitalized_on"])
    result = client.table("aset_data").update(data).eq("id", str(id)).execute()
    if not result.data: raise HTTPException(status_code=404, detail="Not found")
    log_module_update(client, "Data Aset", _admin.get("full_name", "Admin"))
    return result.data[0]

@router.delete("/data/{id}")
def delete_aset_data(id: UUID, _admin: dict = Depends(require_admin)):
    client = get_supabase_admin()
    result = client.table("aset_data").delete().eq("id", str(id)).execute()
    if not result.data: raise HTTPException(status_code=404, detail="Not found")
    log_module_update(client, "Data Aset", _admin.get("full_name", "Admin"))
    return {"message": "Deleted"}

# -------------------------
# UPLOAD & EXPORT EXCEL
# -------------------------
@router.post("/nomor/upload", status_code=status.HTTP_200_OK)
def upload_aset_nomor(file: UploadFile = File(...), _admin: dict = Depends(require_admin)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Harus berupa file Excel (.xlsx / .xls)")
    try:
        content = file.file.read()
        excel_file = pd.ExcelFile(BytesIO(content))
        sheet_name = 'Sheet1' if 'Sheet1' in excel_file.sheet_names else excel_file.sheet_names[0]
        
        # Dynamically find the header row by searching for 'Nomor Aset' or 'Nama Aset'
        df_temp = pd.read_excel(excel_file, sheet_name=sheet_name, header=None, nrows=20)
        header_row = 0
        for i, row in df_temp.iterrows():
            if any(str(x).strip().lower() == 'nomor aset' or str(x).strip().lower() == 'nama aset' for x in row.values):
                header_row = i
                break
                
        df = pd.read_excel(excel_file, sheet_name=sheet_name, header=header_row)
        df.columns = [str(c).strip() for c in df.columns]

        # Map DataFrame columns to DB columns based on expected types
        # This mapping covers the specific Excel file provided by the user
        insert_data = []
        for _, row in df.iterrows():
            def get_val(col_names):
                for name in col_names:
                    if name in df.columns:
                        val = row[name]
                        return val if pd.notna(val) else None
                return None

            nama = get_val(["Nama Aset", "nama_aset", "Deskripsi"])
            nomor = get_val(["Nomor Aset", "nomor_aset"])
            kat = get_val(["Kategori", "kategori"])
            
            # Abaikan baris yang pada dasarnya kosong
            if not nama and not nomor and not kat:
                continue
                
            def clean_str(val):
                if pd.isna(val) or val is None: return None
                s = str(val).strip()
                if s.endswith('.0'):
                    s = s[:-2]
                return s if s != "nan" else None

            item = {
                "nomor_aset": clean_str(nomor),
                "sub_nomor": clean_str(get_val(["Sub Nomor", "sub_nomor", "Sub#"])),
                "nama_aset": clean_str(nama) if clean_str(nama) else "Aset Baru",
                "kategori": clean_str(kat),
                "satuan": clean_str(get_val(["Satuan", "satuan", "Sat"])),
                "lokasi": clean_str(get_val(["Lokasi", "lokasi"])),
                "room": clean_str(get_val(["Room", "room"])),
                "tahun": int(get_val(["Tahun", "tahun"])) if get_val(["Tahun", "tahun"]) else 2026,
                "bulan": int(get_val(["Bulan", "bulan"])) if get_val(["Bulan", "bulan"]) else 1,
                "user_name": str(get_val(["User Name", "user_name", "User"])) if get_val(["User Name", "user_name", "User"]) else None,
                "vendor": str(get_val(["Vendor", "vendor"])) if get_val(["Vendor", "vendor"]) else None,
                "nilai": float(get_val(["Nilai", "nilai"])) if get_val(["Nilai", "nilai"]) else 0.0,
                "keterangan": str(get_val(["Keterangan", "keterangan"])) if get_val(["Keterangan", "keterangan"]) else None,
            }
            insert_data.append(item)

        if not insert_data:
            raise HTTPException(status_code=400, detail="Tidak ada data ditemukan")

        client = get_supabase_admin()
        client.table("aset_nomor").delete().neq("id", "00000000-0000-0000-0000-000000000000").execute()
        
        chunk_size = 100
        for i in range(0, len(insert_data), chunk_size):
            client.table("aset_nomor").insert(insert_data[i:i + chunk_size]).execute()
            
        log_module_update(client, "Nomor Aset", _admin.get("full_name", "Admin"))
        return {"message": f"Berhasil upload {len(insert_data)} data nomor aset.", "count": len(insert_data)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/nomor/export")
def export_aset_nomor(_user: dict = Depends(get_current_user)):
    client = get_supabase_admin()
    
    all_data = []
    chunk_size = 1000
    start = 0
    while True:
        result = client.table("aset_nomor").select("*").order("created_at", desc=False).range(start, start + chunk_size - 1).execute()
        data = result.data
        if not data: break
        all_data.extend(data)
        if len(data) < chunk_size: break
        start += chunk_size
        
    df = pd.DataFrame(all_data)
    
    if not df.empty:
        columns_mapping = {
            'nomor_aset': 'Nomor Aset',
            'sub_nomor': 'Sub#',
            'kategori': 'Kategori',
            'nama_aset': 'Nama Aset',
            'satuan': 'Sat',
            'lokasi': 'Lokasi',
            'nilai': 'Nilai',
            'room': 'Room',
            'tahun': 'Tahun',
            'user_name': 'User',
            'vendor': 'Vendor',
            'keterangan': 'Keterangan'
        }
        
        available_cols = [c for c in columns_mapping.keys() if c in df.columns]
        df = df[available_cols].rename(columns=columns_mapping)
        df.insert(0, 'No', range(1, len(df) + 1))
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Permintaan Nomor Aset')
        worksheet = writer.sheets['Permintaan Nomor Aset']
        
        from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
        
        header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        header_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        green_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
        
        col_widths = {
            'A': 5, 'B': 18, 'C': 8, 'D': 25, 'E': 50, 
            'F': 8, 'G': 15, 'H': 25, 'I': 15, 'J': 8, 
            'K': 20, 'L': 30, 'M': 15
        }
        for col_letter, width in col_widths.items():
            worksheet.column_dimensions[col_letter].width = width
            
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column), start=2):
            is_selesai = False
            keterangan_cell = row[12] if len(row) > 12 else None
            if keterangan_cell and keterangan_cell.value and str(keterangan_cell.value).strip().lower() == 'selesai':
                is_selesai = True
                
            for col_idx, cell in enumerate(row):
                cell.border = thin_border
                
                if col_idx in [0, 2, 5, 8, 9]: 
                    cell.alignment = center_align
                elif col_idx == 7: 
                    cell.number_format = '_("Rp"* #,##0_);_("Rp"* \\(#,##0\\);_("Rp"* "-"_);_(@_)'
                else:
                    cell.alignment = left_align
                    
                if is_selesai:
                    cell.fill = green_fill

    output.seek(0)
    
    headers = {
        'Content-Disposition': 'attachment; filename="Permintaan_Nomor_Aset.xlsx"'
    }
    return Response(content=output.read(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

@router.post("/laporan/upload", status_code=status.HTTP_200_OK)
def upload_aset_laporan(file: UploadFile = File(...), _admin: dict = Depends(require_admin)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Harus berupa file Excel (.xlsx / .xls)")
    try:
        content = file.file.read()
        excel_file = pd.ExcelFile(BytesIO(content))
        sheet_name = 'Sheet1' if 'Sheet1' in excel_file.sheet_names else excel_file.sheet_names[0]
        
        # Dynamically find the header row by searching for 'Asset', 'Asset Number' or 'Asset Description'
        df_temp = pd.read_excel(excel_file, sheet_name=sheet_name, header=None, nrows=20)
        header_row = 0
        for i, row in df_temp.iterrows():
            if any(str(x).strip().lower() in ['asset', 'asset number', 'asset description'] for x in row.values):
                header_row = i
                break
                
        df = pd.read_excel(excel_file, sheet_name=sheet_name, header=header_row)
        df.columns = [str(c).strip() for c in df.columns]

        insert_data = []
        for _, row in df.iterrows():
            def get_val(col_names):
                for name in col_names:
                    if name in df.columns:
                        val = row[name]
                        return val if pd.notna(val) else None
                return None

            def get_date(val):
                if pd.isna(val) or val == "": return None
                if hasattr(val, "date"): return str(val.date())
                
                s = str(val).split()[0]
                if s == '' or s == '—': return None
                try:
                    dt = pd.to_datetime(s, dayfirst=True)
                    if pd.notna(dt):
                        return dt.strftime('%Y-%m-%d')
                except:
                    pass
                return s

            def clean_str(val):
                if pd.isna(val) or val is None: return None
                s = str(val).strip()
                if s.endswith('.0'):
                    s = s[:-2]
                return s if s != "nan" and s != "" else None

            def get_num(val):
                if pd.isna(val) or val is None or str(val).strip() == "": return 0
                try:
                    return float(val)
                except:
                    return 0

            item = {
                "deskripsi": clean_str(get_val(["Deskripsi", "deskripsi"])),
                "asset_number": clean_str(get_val(["Asset", "Asset Number", "asset_number"])),
                "sub_number": clean_str(get_val(["Sub number", "Sub Number", "sub_number"])),
                "capitalized_on": get_date(get_val(["Capitalized on", "Capitalized On", "capitalized_on"])),
                "asset_description": clean_str(get_val(["Asset description", "Asset Description", "asset_description"])),
                "acquis_val": get_num(get_val(["Acquis.val.", "acquis_val"])),
                "accum_dep": get_num(get_val(["Accum.dep.", "accum_dep"])),
                "book_val": get_num(get_val(["Book val.", "book_val"])),
                "currency": clean_str(get_val(["Currency", "currency"])) or "IDR",
                "useful_life": clean_str(get_val(["Useful", "Useful Life", "useful_life"])),
                "location_code": clean_str(get_val(["Location", "Location Code", "location_code"])),
                "lokasi": clean_str(get_val(["Lokasi", "lokasi"])),
                "room": clean_str(get_val(["Room", "room"])),
            }
            if item["asset_number"] or item["asset_description"]:
                insert_data.append(item)

        if not insert_data:
            raise HTTPException(status_code=400, detail="Tidak ada data ditemukan")

        client = get_supabase_admin()
        client.table("aset_laporan_aktiva").delete().neq("id", "00000000-0000-0000-0000-000000000000").execute()
        
        chunk_size = 100
        for i in range(0, len(insert_data), chunk_size):
            client.table("aset_laporan_aktiva").insert(insert_data[i:i + chunk_size]).execute()
            
        log_module_update(client, "Laporan Aktiva Tetap", _admin.get("full_name", "Admin"))
        return {"message": f"Berhasil upload {len(insert_data)} data laporan aktiva.", "count": len(insert_data)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/laporan/export")
def export_aset_laporan(_user: dict = Depends(get_current_user)):
    client = get_supabase_admin()
    
    all_data = []
    chunk_size = 1000
    start = 0
    while True:
        result = client.table("aset_laporan_aktiva").select("*").order("created_at", desc=False).range(start, start + chunk_size - 1).execute()
        data = result.data
        if not data: break
        all_data.extend(data)
        if len(data) < chunk_size: break
        start += chunk_size
        
    df = pd.DataFrame(all_data)
    
    if not df.empty:
        columns_mapping = {
            'deskripsi': 'Deskripsi',
            'asset_number': 'Asset',
            'sub_number': 'Sub number',
            'capitalized_on': 'Capitalized on',
            'asset_description': 'Asset description',
            'acquis_val': 'Acquis.val.',
            'accum_dep': 'Accum.dep.',
            'book_val': 'Book val.',
            'currency': 'Currency',
            'useful_life': 'Useful',
            'location_code': 'Location',
            'lokasi': 'Lokasi',
            'room': 'Room'
        }
        
        available_cols = [c for c in columns_mapping.keys() if c in df.columns]
        df = df[available_cols].rename(columns=columns_mapping)
        df.insert(0, 'No', range(1, len(df) + 1))
        
        if 'Capitalized on' in df.columns:
            df['Capitalized on'] = pd.to_datetime(df['Capitalized on'], errors='coerce').dt.strftime('%d/%m/%y').fillna('')
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Laporan Aktiva Tetap')
        worksheet = writer.sheets['Laporan Aktiva Tetap']
        
        if not df.empty:
            worksheet.auto_filter.ref = worksheet.dimensions
            
        from openpyxl.styles import PatternFill, Border, Side, Alignment
        
        header_fill = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        col_widths = {
            'A': 5,  'B': 15, 'C': 15, 'D': 12, 'E': 15, 
            'F': 45, 'G': 18, 'H': 18, 'I': 18, 'J': 10, 
            'K': 10, 'L': 12, 'M': 20, 'N': 15
        }
        for col_letter, width in col_widths.items():
            worksheet.column_dimensions[col_letter].width = width
            
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column), start=2):
            for col_idx, cell in enumerate(row):
                cell.border = thin_border
                
                if col_idx in [0, 3, 4, 9, 10, 11]: 
                    cell.alignment = center_align
                elif col_idx in [6, 7, 8]: 
                    cell.number_format = '#,##0'
                else:
                    cell.alignment = left_align

    output.seek(0)
    
    headers = {
        'Content-Disposition': 'attachment; filename="Laporan_Aktiva_Tetap.xlsx"'
    }
    return Response(content=output.read(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

# -------------------------
# DASHBOARD ENDPOINT
# -------------------------
@router.get("/dashboard")
def get_dashboard_metrics(tahun: int = Query(2026), _user: dict = Depends(get_current_user)):
    client = get_supabase_admin()
    # Fetch all data using pagination since Supabase limits to 1000 per request
    laporan_data = []
    start = 0
    while True:
        res = client.table("aset_laporan_aktiva").select("*").range(start, start + 999).execute()
        if not res.data: break
        laporan_data.extend(res.data)
        if len(res.data) < 1000: break
        start += 1000

    nomor_data = []
    start = 0
    while True:
        res = client.table("aset_nomor").select("*").eq("tahun", tahun).range(start, start + 999).execute()
        if not res.data: break
        nomor_data.extend(res.data)
        if len(res.data) < 1000: break
        start += 1000
    
    return {
        "laporan": laporan_data,
        "nomor": nomor_data
    }
