from fastapi import APIRouter, Depends, HTTPException, Query, status, UploadFile, File
import openpyxl
from io import BytesIO
from typing import Optional
from uuid import UUID

from ..core.database import get_supabase_admin
from ..core.security import get_current_user, require_admin
from ..models.asset import AssetCreate, AssetUpdate, AssetResponse

router = APIRouter(prefix="/assets", tags=["Aset"])

_TABLE = "capex_assets"


@router.get("", response_model=list[AssetResponse])
def list_assets(
    category: Optional[str] = None,
    lokasi: Optional[str] = None,
    search: Optional[str] = None,
    _user: dict = Depends(get_current_user),
):
    client = get_supabase_admin()
    query = client.table(_TABLE).select("*").order("tanggal_po", desc=True)

    if category is not None and isinstance(category, str):
        query = query.eq("category", category)
    if lokasi is not None and isinstance(lokasi, str):
        query = query.ilike("lokasi", f"%{lokasi}%")
    if search is not None and isinstance(search, str):
        query = query.ilike("asset_description", f"%{search}%")

    result = query.execute()
    return result.data


@router.get("/{asset_id}", response_model=AssetResponse)
def get_asset(
    asset_id: UUID,
    _user: dict = Depends(get_current_user),
):
    client = get_supabase_admin()
    result = client.table(_TABLE).select("*").eq("id", str(asset_id)).single().execute()
    if not result.data:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Data aset tidak ditemukan.")
    return result.data


@router.post("", response_model=AssetResponse, status_code=status.HTTP_201_CREATED)
def create_asset(
    payload: AssetCreate,
    _admin: dict = Depends(require_admin),
):
    client = get_supabase_admin()
    data = payload.model_dump()
    for date_field in ("tanggal_po", "capitalized_on", "kajian_tanggal"):
        if data.get(date_field):
            data[date_field] = str(data[date_field])
    result = client.table(_TABLE).insert(data).execute()
    return result.data[0]


@router.put("/{asset_id}", response_model=AssetResponse)
def update_asset(
    asset_id: UUID,
    payload: AssetUpdate,
    _admin: dict = Depends(require_admin),
):
    client = get_supabase_admin()
    update_data = payload.model_dump(exclude_none=True)
    if not update_data:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Tidak ada field yang diupdate.")

    for date_field in ("tanggal_po", "capitalized_on", "kajian_tanggal"):
        if date_field in update_data:
            update_data[date_field] = str(update_data[date_field])

    result = client.table(_TABLE).update(update_data).eq("id", str(asset_id)).execute()
    if not result.data:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Data aset tidak ditemukan.")
    return result.data[0]


@router.delete("/{asset_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_asset(
    asset_id: UUID,
    _admin: dict = Depends(require_admin),
):
    client = get_supabase_admin()
    result = client.table(_TABLE).delete().eq("id", str(asset_id)).execute()
    if not result.data:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Data aset tidak ditemukan.")

@router.post("/upload", status_code=status.HTTP_200_OK)
def upload_assets(
    file: UploadFile = File(...),
    _admin: dict = Depends(require_admin),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Harus berupa file Excel (.xlsx / .xls)")
    
    try:
        content = file.file.read()
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        sheet_names = wb.sheetnames
        sheet = wb["Per Kategori"] if "Per Kategori" in sheet_names else wb.active
        
        insert_data = []
        current_kategori = None
        
        for row in sheet.iter_rows(min_row=5, values_only=True): # start from row 5 because header is at 4
            # Check for category header in 'Per Kategori' tab
            if row[0] is None and row[1] is not None and str(row[1]).strip() != "":
                if row[10] is None or str(row[10]).strip() == "":
                    current_kategori = str(row[1]).strip()
                    continue
            
            if not row[10]: # Asset description is empty, probably end of table or empty row
                continue
            
            def get_val(val, default=None):
                return val if val is not None and str(val).strip() != "" else default
                
            def get_date(val):
                if not val: return None
                if hasattr(val, "date"): return str(val.date())
                return str(val).split()[0]
                
            def get_int(val):
                if not val: return 0
                try:
                    return int(float(val))
                except:
                    return 0

            item = {
                "kajian_no": str(get_val(row[1])) if get_val(row[1]) else None,
                "kajian_tanggal": get_date(row[2]),
                "kajian_perihal": str(get_val(row[3])) if get_val(row[3]) else None,
                "no_po": str(get_val(row[4])) if get_val(row[4]) else None,
                "tanggal_po": get_date(row[5]),
                "no_asset": str(get_val(row[6])) if get_val(row[6]) else None,
                "sub_number": str(get_val(row[7])) if get_val(row[7]) else None,
                "category": str(get_val(row[8])) if get_val(row[8]) else None,
                "capitalized_on": get_date(row[9]),
                "asset_description": str(get_val(row[10])),
                "acquis_val": get_int(row[11]),
                "accum_dep": get_int(row[12]),
                "book_val": get_int(row[13]),
                "currency": str(get_val(row[14], "IDR"))[:3],
                "location_code": str(get_val(row[15])) if get_val(row[15]) else None,
                "lokasi": str(get_val(row[16])) if get_val(row[16]) else None,
                "room": str(get_val(row[17])) if get_val(row[17]) else None,
                "keterangan": str(get_val(row[18])) if get_val(row[18]) else None,
                "kategori_aset": current_kategori,
            }
            insert_data.append(item)
            
        if not insert_data:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Tidak ada data yang valid ditemukan di baris 6 ke bawah.")

        client = get_supabase_admin()
        # Truncate existing data (Atomic delete)
        client.table(_TABLE).delete().neq("id", "00000000-0000-0000-0000-000000000000").execute()
        
        # Bulk insert new data
        chunk_size = 100
        for i in range(0, len(insert_data), chunk_size):
            chunk = insert_data[i:i + chunk_size]
            client.table(_TABLE).insert(chunk).execute()
            
        return {"message": f"Berhasil mengupload {len(insert_data)} data aset.", "count": len(insert_data)}
        
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Gagal memproses file: {str(e)}")
